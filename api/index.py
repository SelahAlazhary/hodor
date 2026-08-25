# -*- coding: utf-8 -*-
"""
═══════════════════════════════════════════════════════════════
  سلاح الأزهري — خدمة التقارير والمهام التلقائية (Python / FastAPI)
═══════════════════════════════════════════════════════════════
تعمل بجانب التطبيق ولا تعطّله: التطبيق يظل يسجّل الحضور مباشرة
(ويعمل بدون إنترنت)، وهذه الخدمة تتولّى الأعمال الثقيلة:

  • تقارير Excel احترافية (شهرية / يومية / كشف رواتب)
  • تسجيل الغياب تلقائياً آخر كل يوم عمل
  • ملخّصات جاهزة بصيغة JSON

المصدر: قاعدة Firebase Realtime Database عبر واجهة REST.
"""

import io
import json
import os
import urllib.request
import urllib.error
import urllib.parse
from datetime import datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from fastapi import FastAPI, Query
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════ الإعدادات ═══════════
DB_URL = os.environ.get(
    "RTDB_URL", "https://hodor-abfe0-default-rtdb.firebaseio.com"
).rstrip("/")
# توقيت القاهرة — نعتمد على حزمة tzdata المرفقة، ومع فشلها نستخدم +03:00
try:
    TZ = ZoneInfo("Africa/Cairo")
    TZ_SOURCE = "Africa/Cairo"
except Exception:                                    # pragma: no cover
    TZ = timezone(timedelta(hours=3))
    TZ_SOURCE = "UTC+03:00 (احتياطي)"

AR_DAYS = ["الأحد", "الإثنين", "الثلاثاء", "الأربعاء", "الخميس", "الجمعة", "السبت"]
AR_MONTHS = ["يناير", "فبراير", "مارس", "أبريل", "مايو", "يونيو",
             "يوليو", "أغسطس", "سبتمبر", "أكتوبر", "نوفمبر", "ديسمبر"]

app = FastAPI(
    title="نظام حضور سلاح الأزهري",
    description="خدمة التقارير والمهام التلقائية",
    docs_url="/api/docs",
    openapi_url="/api/openapi.json",
)


# ═══════════ الوصول إلى قاعدة البيانات ═══════════
def db_get(path: str, default=None):
    """قراءة مسار من قاعدة البيانات اللحظية."""
    url = f"{DB_URL}/{path.strip('/')}.json"
    try:
        with urllib.request.urlopen(url, timeout=20) as r:
            data = json.loads(r.read().decode("utf-8"))
            return default if data is None else data
    except Exception:
        return default


def db_write(path: str, payload, method: str = "PATCH"):
    """كتابة إلى قاعدة البيانات اللحظية."""
    url = f"{DB_URL}/{path.strip('/')}.json"
    body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        url, data=body, method=method,
        headers={"Content-Type": "application/json; charset=utf-8"},
    )
    with urllib.request.urlopen(req, timeout=20) as r:
        return json.loads(r.read().decode("utf-8") or "null")


# ═══════════ أدوات الوقت ═══════════
def now_cairo() -> datetime:
    return datetime.now(TZ)


def date_key(d: datetime = None) -> str:
    return (d or now_cairo()).strftime("%Y-%m-%d")


def month_range(month: str):
    """يعيد (أول يوم، آخر يوم) لشهر بصيغة YYYY-MM."""
    y, m = (int(x) for x in month.split("-"))
    start = datetime(y, m, 1, tzinfo=TZ)
    end = datetime(y + (m == 12), (m % 12) + 1, 1, tzinfo=TZ) - timedelta(days=1)
    return start.strftime("%Y-%m-%d"), end.strftime("%Y-%m-%d")


def date_ar(dk: str) -> str:
    d = datetime.strptime(dk, "%Y-%m-%d")
    return f"{d.day} {AR_MONTHS[d.month - 1]} {d.year}"


def day_ar(dk: str) -> str:
    d = datetime.strptime(dk, "%Y-%m-%d")
    return AR_DAYS[(d.weekday() + 1) % 7]        # الاثنين=0 → الأحد=0


def weekday_num(dk: str) -> int:
    """رقم اليوم بنفس ترتيب JavaScript (الأحد=0 … السبت=6)."""
    return (datetime.strptime(dk, "%Y-%m-%d").weekday() + 1) % 7


def time_ar(iso: str) -> str:
    """يحوّل لحظة ISO إلى ساعة بتوقيت القاهرة بنظام 24 ساعة (17:05)."""
    if not iso:
        return "—"
    try:
        d = datetime.fromisoformat(str(iso).replace("Z", "+00:00")).astimezone(TZ)
    except Exception:
        return "—"
    return f"{d.hour:02d}:{d.minute:02d}"


def hhmm_to_min(t):
    try:
        h, m = str(t).split(":")
        return int(h) * 60 + int(m)
    except Exception:
        return None


def min_to_hours(m) -> float:
    return round((float(m or 0)) / 60, 2)


# ═══════════ منطق الدوام (مطابق تماماً لواجهة التطبيق) ═══════════
def required_min(rec: dict, emp: dict, settings: dict) -> int:
    """الدقائق المطلوبة في اليوم لهذا الموظف."""
    a = hhmm_to_min((rec or {}).get("expectedStart") or (emp or {}).get("workStart") or settings.get("workStart"))
    b = hhmm_to_min((rec or {}).get("expectedEnd") or (emp or {}).get("workEnd") or settings.get("workEnd"))
    if a is not None and b is not None:
        d = b - a
        if d < 0:
            d += 1440
        if d > 0:
            return d
    return int(round(float(settings.get("dailyHours") or 8) * 60))


def overtime_min(rec: dict, emp: dict, settings: dict) -> int:
    if not rec or not rec.get("checkOut"):
        return 0
    if isinstance(rec.get("overtimeMin"), (int, float)):
        return max(0, int(rec["overtimeMin"]))
    return max(0, int(rec.get("workedMin") or 0) - required_min(rec, emp, settings))


def summarize(rows, emp, settings) -> dict:
    s = {"days": 0, "absent": 0, "late": 0, "minutes": 0, "overtime": 0, "open": 0}
    for r in rows:
        if r.get("status") == "absent":
            s["absent"] += 1
            continue
        if r.get("checkIn"):
            s["days"] += 1
            if r.get("status") == "late":
                s["late"] += 1
            if r.get("checkOut"):
                s["minutes"] += int(r.get("workedMin") or 0)
                s["overtime"] += overtime_min(r, emp, settings)
            else:
                s["open"] += 1
    s["hours"] = min_to_hours(s["minutes"])
    s["overtimeHours"] = min_to_hours(s["overtime"])
    s["avgHours"] = min_to_hours(s["minutes"] / s["days"]) if s["days"] else 0
    return s


def load_context(month: str = None):
    """يحمّل الإعدادات والموظفين وسجلات الشهر دفعة واحدة."""
    settings = db_get("settings/global", {}) or {}
    employees = db_get("employees", {}) or {}
    emps = [dict(v, id=k) for k, v in employees.items()]
    emps.sort(key=lambda e: str(e.get("name", "")))

    rows_by_emp = {}
    if month:
        start, end = month_range(month)
        attendance = db_get("attendance", {}) or {}
        for dk, per_emp in attendance.items():
            if not (start <= dk <= end) or not isinstance(per_emp, dict):
                continue
            for eid, rec in per_emp.items():
                rows_by_emp.setdefault(eid, []).append(dict(rec, date=dk, empId=eid))
        for v in rows_by_emp.values():
            v.sort(key=lambda r: r["date"])
    return settings, emps, rows_by_emp


# ═══════════ تنسيق ملفات Excel ═══════════
GREEN = "0B6B3A"
GOLD = "C9A227"
LIGHT = "E3F3EA"
THIN = Side(style="thin", color="C9D6CE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


_FIRST = "_az_first_sheet_used"


def new_sheet(wb, title: str, headers: list, widths: list):
    """ينشئ ورقة منسّقة من اليمين لليسار."""
    if not getattr(wb, _FIRST, False):
        ws = wb.active
        setattr(wb, _FIRST, True)
    else:
        ws = wb.create_sheet()
    ws.title = title
    ws.sheet_view.rightToLeft = True
    ws.append(headers)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = Font(name="Cairo", bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    return ws


def style_body(ws, first_row=2, bold_last=False):
    for row in ws.iter_rows(min_row=first_row, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name="Cairo", size=10.5)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = BORDER
    if bold_last:
        for cell in ws[ws.max_row]:
            cell.font = Font(name="Cairo", bold=True, size=11)
            cell.fill = PatternFill("solid", fgColor=LIGHT)


def xlsx_response(wb, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    quoted = urllib.parse.quote(filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted}"},
    )


# ═══════════════════════ نقاط الخدمة ═══════════════════════
@app.get("/api/health")
def health():
    """فحص الخدمة والوقت الرسمي."""
    settings = db_get("settings/global", {}) or {}
    emps = db_get("employees", {}) or {}
    n = now_cairo()
    return {
        "ok": True,
        "service": "خدمة تقارير سلاح الأزهري",
        "serverTimeCairo": n.strftime("%Y-%m-%d %H:%M:%S"),
        "day": day_ar(date_key(n)),
        "employees": len(emps),
        "workHours": f"{settings.get('workStart', '—')} → {settings.get('workEnd', '—')}",
        "timezone": TZ_SOURCE,
        "database": DB_URL,
    }


@app.get("/api/summary")
def summary(month: str = Query(default=None, description="YYYY-MM")):
    """ملخّص شهري بصيغة JSON."""
    month = month or now_cairo().strftime("%Y-%m")
    settings, emps, rows_by_emp = load_context(month)
    out, tot_h, tot_ot = [], 0.0, 0.0
    for e in emps:
        s = summarize(rows_by_emp.get(e["id"], []), e, settings)
        tot_h += s["hours"]
        tot_ot += s["overtimeHours"]
        out.append({
            "name": e.get("name"),
            "shift": f"{e.get('workStart') or settings.get('workStart')} - {e.get('workEnd') or settings.get('workEnd')}",
            **s,
        })
    return {
        "month": month,
        "employees": out,
        "totalHours": round(tot_h, 2),
        "totalOvertimeHours": round(tot_ot, 2),
    }


@app.get("/api/report")
def monthly_report(month: str = Query(default=None, description="YYYY-MM")):
    """تقرير Excel شهري كامل: ملخّص لكل موظف + تفاصيل الأيام."""
    month = month or now_cairo().strftime("%Y-%m")
    settings, emps, rows_by_emp = load_context(month)
    y, m = (int(x) for x in month.split("-"))
    company = settings.get("company") or "سلاح الأزهري"

    wb = Workbook()
    ws = new_sheet(
        wb, "الملخص الشهري",
        ["الموظف", "الوظيفة", "الدوام", "أيام الحضور", "أيام الغياب",
         "مرات التأخير", "إجمالي الساعات", "ساعات إضافية", "متوسط اليوم"],
        [24, 16, 16, 13, 12, 12, 14, 13, 13],
    )
    tot = {"days": 0, "absent": 0, "late": 0, "hours": 0.0, "ot": 0.0}
    for e in emps:
        s = summarize(rows_by_emp.get(e["id"], []), e, settings)
        ws.append([
            e.get("name", ""), e.get("job", "") or "—",
            f"{e.get('workStart') or settings.get('workStart', '')} - {e.get('workEnd') or settings.get('workEnd', '')}",
            s["days"], s["absent"], s["late"], s["hours"], s["overtimeHours"], s["avgHours"],
        ])
        tot["days"] += s["days"]; tot["absent"] += s["absent"]; tot["late"] += s["late"]
        tot["hours"] += s["hours"]; tot["ot"] += s["overtimeHours"]
    ws.append(["الإجمالي", "", "", tot["days"], tot["absent"], tot["late"],
               round(tot["hours"], 2), round(tot["ot"], 2), ""])
    style_body(ws, bold_last=True)

    # ورقة تفاصيل لكل موظف
    det = new_sheet(
        wb, "تفاصيل الأيام",
        ["الموظف", "التاريخ", "اليوم", "الحضور", "الانصراف",
         "الساعات", "إضافي", "الحالة", "ملاحظة"],
        [22, 15, 12, 12, 12, 11, 11, 11, 20],
    )
    status_ar = {"absent": "غياب", "late": "متأخر", "present": "حاضر"}
    for e in emps:
        for r in rows_by_emp.get(e["id"], []):
            ot = overtime_min(r, e, settings)
            det.append([
                e.get("name", ""), date_ar(r["date"]), day_ar(r["date"]),
                time_ar(r.get("checkIn")), time_ar(r.get("checkOut")),
                min_to_hours(r.get("workedMin")) if r.get("checkOut") else "—",
                min_to_hours(ot) if ot else "—",
                status_ar.get(r.get("status"), "—"),
                r.get("note") or "",
            ])
    style_body(det)

    title = f"تقرير-{company}-{AR_MONTHS[m - 1]}-{y}.xlsx"
    return xlsx_response(wb, title)


@app.get("/api/report/daily")
def daily_report(date: str = Query(default=None, description="YYYY-MM-DD")):
    """تقرير Excel ليوم واحد."""
    dk = date or date_key()
    settings = db_get("settings/global", {}) or {}
    employees = db_get("employees", {}) or {}
    day = db_get(f"attendance/{dk}", {}) or {}

    wb = Workbook()
    ws = new_sheet(
        wb, f"حضور {dk}",
        ["الموظف", "الدوام", "الحضور", "الانصراف", "الساعات", "إضافي", "الحالة"],
        [24, 16, 13, 13, 12, 11, 12],
    )
    for eid, e in sorted(employees.items(), key=lambda kv: str(kv[1].get("name", ""))):
        if e.get("active") is False:
            continue
        r = day.get(eid) or {}
        ot = overtime_min(r, e, settings)
        status = ("غياب" if r.get("status") == "absent"
                  else "انصرف" if r.get("checkOut")
                  else "متأخر" if r.get("status") == "late"
                  else "حاضر" if r.get("checkIn") else "لم يسجّل")
        ws.append([
            e.get("name", ""),
            f"{e.get('workStart') or settings.get('workStart', '')} - {e.get('workEnd') or settings.get('workEnd', '')}",
            time_ar(r.get("checkIn")), time_ar(r.get("checkOut")),
            min_to_hours(r.get("workedMin")) if r.get("checkOut") else "—",
            min_to_hours(ot) if ot else "—", status,
        ])
    style_body(ws)
    return xlsx_response(wb, f"حضور-{dk}.xlsx")


@app.get("/api/payroll")
def payroll(
    month: str = Query(default=None, description="YYYY-MM"),
    rate: float = Query(default=0, description="أجر الساعة الافتراضي"),
    ot_multiplier: float = Query(default=1.5, description="مضاعف الساعة الإضافية"),
):
    """كشف رواتب Excel محسوب من الساعات الفعلية والإضافية."""
    month = month or now_cairo().strftime("%Y-%m")
    settings, emps, rows_by_emp = load_context(month)
    y, m = (int(x) for x in month.split("-"))

    wb = Workbook()
    ws = new_sheet(
        wb, "كشف الرواتب",
        ["الموظف", "أجر الساعة", "ساعات عادية", "قيمة العادي",
         "ساعات إضافية", f"قيمة الإضافي (×{ot_multiplier})", "الإجمالي المستحق"],
        [24, 13, 14, 14, 14, 18, 16],
    )
    grand = 0.0
    for e in emps:
        s = summarize(rows_by_emp.get(e["id"], []), e, settings)
        r = float(e.get("hourRate") or rate or 0)
        normal_h = round(max(0.0, s["hours"] - s["overtimeHours"]), 2)
        normal_v = round(normal_h * r, 2)
        ot_v = round(s["overtimeHours"] * r * ot_multiplier, 2)
        total = round(normal_v + ot_v, 2)
        grand += total
        ws.append([e.get("name", ""), r, normal_h, normal_v,
                   s["overtimeHours"], ot_v, total])
    ws.append(["الإجمالي", "", "", "", "", "", round(grand, 2)])
    style_body(ws, bold_last=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in (row[3], row[5], row[6]):
            cell.number_format = "#,##0.00"
    return xlsx_response(wb, f"رواتب-{AR_MONTHS[m - 1]}-{y}.xlsx")


@app.get("/api/auto-absent")
def auto_absent(
    date: str = Query(default=None, description="YYYY-MM-DD (افتراضياً اليوم)"),
    dry: bool = Query(default=False, description="معاينة بدون كتابة"),
):
    """يسجّل غياب كل موظف لم يحضر — يعمل تلقائياً آخر كل يوم عمل عبر Vercel Cron."""
    dk = date or date_key()
    settings = db_get("settings/global", {}) or {}
    employees = db_get("employees", {}) or {}
    day = db_get(f"attendance/{dk}", {}) or {}

    weekend = settings.get("weekend")
    weekend = weekend if isinstance(weekend, list) else [5]      # الجمعة افتراضياً
    if weekday_num(dk) in weekend:
        return {"ok": True, "skipped": "إجازة أسبوعية", "date": dk, "day": day_ar(dk), "marked": 0}

    marked = []
    for eid, e in employees.items():
        if e.get("active") is False or eid in day:
            continue
        marked.append(e.get("name"))
        if dry:
            continue
        now_ms = int(now_cairo().timestamp() * 1000)
        db_write(f"attendance/{dk}/{eid}", {
            "empId": eid, "empName": e.get("name"), "date": dk,
            "checkIn": None, "checkOut": None, "workedMin": 0,
            "status": "absent", "note": "تسجيل تلقائي آخر اليوم",
            "markedBy": "system", "createdAt": now_ms, "updatedAt": now_ms,
        })
        db_write("events", {
            "type": "abs", "empId": eid, "empName": e.get("name"), "date": dk,
            "at": now_cairo().isoformat(), "note": "تسجيل تلقائي",
            "createdAt": now_ms,
        }, method="POST")

    return {"ok": True, "date": dk, "day": day_ar(dk),
            "marked": len(marked), "employees": marked, "dry": dry}


@app.get("/api")
def index():
    return JSONResponse({
        "service": "خدمة تقارير سلاح الأزهري",
        "endpoints": {
            "/api/health": "حالة الخدمة والوقت الرسمي",
            "/api/summary?month=YYYY-MM": "ملخّص شهري JSON",
            "/api/report?month=YYYY-MM": "تقرير Excel شهري",
            "/api/report/daily?date=YYYY-MM-DD": "تقرير Excel يومي",
            "/api/payroll?month=YYYY-MM&rate=50": "كشف رواتب Excel",
            "/api/auto-absent": "تسجيل الغائبين تلقائياً",
        },
    })
